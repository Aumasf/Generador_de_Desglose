"""
match_utils.py

Se encarga de:
- Leer el archivo match.xlsx (tabla de coincidencias).
- Buscar la mejor coincidencia (>= 80%) entre la descripción del ítem y la tabla.
- Completar 4 textos para el PDF:
    1) Equipos a utilizar  -> desde columna "Herramientas"
    2) Mano de obra       -> texto fijo (o vacío según reglas)
    3) Materiales         -> desde columna "Materiales" (o vacío según reglas)
    4) Transporte         -> texto fijo

IMPORTANTE:
- Este módulo NO modifica cálculos ni coordenadas de números.
- Solo agrega campos de texto a cada fila.

Dependencias:
- Usa openpyxl (ya está en requirements.txt del proyecto).
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
import re
import unicodedata
from typing import Dict, List, Optional, Tuple

from openpyxl import load_workbook


# ==========================================================
# Normalización y tokens (acentos/stopwords)
# ==========================================================

_STOPWORDS = {
    # artículos / preposiciones / conectores comunes
    "el", "la", "los", "las", "un", "una", "unos", "unas",
    "de", "del", "al", "a", "en", "y", "o", "u", "para", "por", "con", "sin",
    "sobre", "segun", "según", "entre", "desde", "hasta", "tipo",
    # palabras poco útiles en descripciones técnicas
    "servicio", "trabajo", "trabajos", "reparacion", "reparación", "mantenimiento",
    "equipo", "equipos", "unidad", "unidades"
}


def _normalize(s: str) -> str:
    s = str(s or "").strip().lower()
    s = unicodedata.normalize("NFD", s)
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")  # quita acentos
    s = re.sub(r"[^a-z0-9\s]+", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def _tokens(s: str) -> List[str]:
    s = _normalize(s)
    toks = [t for t in s.split(" ") if t and t not in _STOPWORDS and len(t) >= 2]
    return toks


def _coverage_similarity(query_tokens: List[str], ref_tokens: List[str]) -> float:
    """
    Similaridad tipo "coverage":
      intersección / tokens_ref

    Esto funciona bien para "descripción larga" vs "descripción base".
    """
    if not ref_tokens:
        return 0.0
    q = set(query_tokens)
    r = set(ref_tokens)
    inter = q.intersection(r)
    return len(inter) / max(len(r), 1)


# ==========================================================
# Reglas de clasificación (material / mano de obra / ambiguo)
# ==========================================================

def _clasificar_item(descripcion: str) -> str:
    """
    Devuelve: "materiales" | "mano_obra" | "ambiguo"

    Reglas (sin acentos y sin mayúsculas):
    - Si empieza con "provision de" o "provicion de" -> materiales
    - Si contiene "mano de obra", "montaje", "desmontaje" (o similares) -> mano de obra
    - Si no, ambiguo
    """
    d = _normalize(descripcion)

    # 1) Provisión -> materiales
    if d.startswith("provision de") or d.startswith("provicion de"):
        return "materiales"

    # 2) Palabras que indican mano de obra
    claves_mo = [
        "mano de obra",
        "montaje",
        "desmontaje",
        "instalacion",
        "colocacion",
        "retiro",
        "reemplazo",
        "mantenimiento",
        "reparacion",
    ]
    for k in claves_mo:
        if k in d:
            return "mano_obra"

    return "ambiguo"


# ==========================================================
# Lectura y búsqueda de coincidencias
# ==========================================================

@dataclass
class MatchRow:
    descripcion_raw: str
    descripcion_norm: str
    tokens: List[str]
    herramientas: str
    materiales: str


def cargar_match_table(path_match_xlsx: Path) -> Tuple[List[MatchRow], MatchRow]:
    """
    Carga match.xlsx y devuelve:
      (filas, fila_default)

    Se espera columnas:
      - Descripcion
      - Herramientas
      - Materiales

    La fila default se identifica por Descripcion == "DEFAULT" (insensible a mayúsculas).
    """
    wb = load_workbook(path_match_xlsx, data_only=True)
    ws = wb.active

    # Encabezados en la primera fila
    headers = {}
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col).value
        if val is None:
            continue
        headers[_normalize(val)] = col

    col_desc = headers.get("descripcion")
    col_herr = headers.get("herramientas")
    col_mat = headers.get("materiales")

    if not col_desc or not col_herr or not col_mat:
        raise ValueError("match.xlsx debe contener columnas: Descripcion, Herramientas, Materiales")

    rows: List[MatchRow] = []
    default_row: Optional[MatchRow] = None

    for r in range(2, ws.max_row + 1):
        desc_raw = str(ws.cell(row=r, column=col_desc).value or "").strip()
        herr = str(ws.cell(row=r, column=col_herr).value or "").strip()
        mat = str(ws.cell(row=r, column=col_mat).value or "").strip()

        desc_norm = _normalize(desc_raw)

        mr = MatchRow(
            descripcion_raw=desc_raw,
            descripcion_norm=desc_norm,
            tokens=_tokens(desc_raw),
            herramientas=herr,
            materiales=mat,
        )

        if desc_norm == "default":
            default_row = mr
        else:
            rows.append(mr)

    wb.close()

    if default_row is None:
        # último recurso: default sintético
        default_row = MatchRow(
            descripcion_raw="DEFAULT",
            descripcion_norm="default",
            tokens=[],
            herramientas="Herramientas de mano",
            materiales="Insumos y materiales",
        )

    return rows, default_row


def buscar_mejor_match(descripcion_item: str, rows: List[MatchRow], default_row: MatchRow, umbral: float = 0.80) -> MatchRow:
    """
    Devuelve la mejor coincidencia. Si ninguna supera el umbral -> default.
    """
    q_tokens = _tokens(descripcion_item)
    if not q_tokens:
        return default_row

    mejor = None
    mejor_score = 0.0
    for r in rows:
        score = _coverage_similarity(q_tokens, r.tokens)
        if score > mejor_score:
            mejor_score = score
            mejor = r

    if mejor is None or mejor_score < umbral:
        return default_row

    return mejor


# ==========================================================
# API principal: aplicar match a filas
# ==========================================================

def aplicar_match_a_filas(filas: List[Dict], path_match_xlsx: Path) -> List[Dict]:
    """
    Recibe las filas que ya vienen del Excel principal (con item/descripcion/etc)
    y devuelve una NUEVA lista con campos extra:

      - texto_equipos
      - texto_mano_obra
      - texto_materiales
      - texto_transporte

    Reglas:
      - Transporte SIEMPRE: "Transporte terrestre"
      - Equipos SIEMPRE: desde match.xlsx (si vacío -> "Herramientas de mano")
      - Mano de obra:
          * si item == materiales -> vacío
          * si mano_obra o ambiguo -> "Supervisor, técnicos oficiales y técnicos ayudantes"
      - Materiales:
          * si item == mano_obra -> vacío
          * si materiales o ambiguo -> desde match.xlsx (si vacío -> "Insumos y materiales")
    """
    rows, default_row = cargar_match_table(path_match_xlsx)

    out: List[Dict] = []
    for fila in filas:
        desc = str(fila.get("descripcion", "") or "").strip()

        mr = buscar_mejor_match(desc, rows, default_row, umbral=0.80)

        tipo = _clasificar_item(desc)

        # Siempre existen:
        texto_transporte = "Transporte terrestre"
        texto_equipos = (mr.herramientas or "").strip() or "Herramientas de mano"

        # Mano de obra / Materiales según reglas
        texto_mano_obra = ""
        texto_materiales = ""

        if tipo == "materiales":
            texto_materiales = (mr.materiales or "").strip() or "Insumos y materiales"
        elif tipo == "mano_obra":
            texto_mano_obra = "Supervisor, técnicos oficiales y técnicos ayudantes"
        else:  # ambiguo
            texto_mano_obra = "Supervisor, técnicos oficiales y técnicos ayudantes"
            texto_materiales = (mr.materiales or "").strip() or "Insumos y materiales"

        fila2 = dict(fila)
        fila2["texto_equipos"] = texto_equipos
        fila2["texto_mano_obra"] = texto_mano_obra
        fila2["texto_materiales"] = texto_materiales
        fila2["texto_transporte"] = texto_transporte

        # NUEVO: guardamos el tipo de ítem para que la lógica de costos
        # (D/E/F/A/B) pueda repartir correctamente el CDT.
        # Valores posibles: "materiales", "mano_obra", "ambiguo".
        fila2["tipo_item"] = tipo

        out.append(fila2)

    return out
