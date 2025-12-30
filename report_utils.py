import csv
import re
import unicodedata
from pathlib import Path
from openpyxl import load_workbook


def normalizar(texto: str) -> str:
    """
    Quita acentos, pasa a minúsculas y elimina espacios extra.
    """
    texto = "" if texto is None else str(texto)
    texto = texto.strip().lower()
    texto = unicodedata.normalize("NFD", texto)
    texto = "".join(c for c in texto if unicodedata.category(c) != "Mn")
    texto = re.sub(r"\s+", " ", texto)
    return texto


def to_number(valor):
    """
    Convierte strings con separadores de miles/decimales a float.
    Soporta formatos típicos de planillas en PY (puntos miles, coma decimal).
    """
    if valor is None:
        return None
    if isinstance(valor, (int, float)):
        return float(valor)

    s = str(valor).strip()
    if s == "":
        return None

    s_norm = normalizar(s)
    if s_norm in ("_no_aplica_", "no aplica", "n/a", "na"):
        return None

    s = s.replace(" ", "")
    # quitar símbolos y letras, conservar dígitos y separadores y signo
    s = re.sub(r"[^0-9,.\-]", "", s)

    if s == "" or s == "-" or s == "." or s == ",":
        return None

    # Si hay coma y punto, asumimos . miles y , decimal
    if "," in s and "." in s:
        s = s.replace(".", "")
        s = s.replace(",", ".")
    else:
        # muchos puntos => miles
        if s.count(".") > 1:
            s = s.replace(".", "")
        # coma única => decimal
        if s.count(",") > 1:
            s = s.replace(",", "")
        else:
            s = s.replace(",", ".")

    try:
        return float(s)
    except Exception:
        return None


def _item_to_int(valor):
    if valor is None:
        return None
    if isinstance(valor, (int, float)):
        try:
            return int(valor)
        except Exception:
            return None
    s = str(valor).strip()
    m = re.search(r"-?\d+", s)
    return int(m.group(0)) if m else None


def leer_reporte(path: Path) -> dict:
    """
    Lee un reporte (CSV o Excel) y devuelve un dict {item_int: precio_unit_referencial}.
    Intenta identificar columnas por nombre, ignorando mayúsculas/acentos.
    """
    path = Path(path)
    ext = path.suffix.lower()

    if ext == ".csv":
        return _leer_reporte_csv(path)
    elif ext in (".xlsx", ".xlsm", ".xls"):
        return _leer_reporte_xlsx(path)
    else:
        raise ValueError("Formato de reporte no soportado. Use .csv o .xlsx/.xlsm/.xls")


def _detectar_campos(headers_norm):
    # candidatos item
    item_field = None
    price_field = None

    for h in headers_norm:
        if item_field is None and (h in ("numero", "nro", "item", "ítem") or h.startswith("item") or "numero" in h):
            item_field = h

    # precio unitario referencial/estimado
    for h in headers_norm:
        if price_field is None and ("precio" in h and ("unit" in h or "unitario" in h)):
            # si existe algo tipo estimado/referencial, priorizar
            if any(k in h for k in ("estim", "refer", "ref")):
                price_field = h

    # fallback: el primero que cumpla precio+unitario
    if price_field is None:
        for h in headers_norm:
            if "precio" in h and ("unit" in h or "unitario" in h):
                price_field = h
                break

    return item_field, price_field


def _leer_reporte_csv(path: Path) -> dict:
    with open(path, "r", encoding="utf-8-sig", errors="replace", newline="") as f:
        sample = f.read(4096)
        f.seek(0)

        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=";,")
        except Exception:
            # por defecto en DNCP suele ser ';'
            class _D: delimiter = ";"
            dialect = _D()

        reader = csv.reader(f, delimiter=dialect.delimiter)
        headers = next(reader, [])
        headers_norm = [normalizar(h) for h in headers]

        item_h, price_h = _detectar_campos(headers_norm)
        if item_h is None or price_h is None:
            raise ValueError("No se pudieron identificar columnas de ítem y precio unitario en el reporte.")

        idx_item = headers_norm.index(item_h)
        idx_price = headers_norm.index(price_h)

        out = {}
        for row in reader:
            if not row or len(row) <= max(idx_item, idx_price):
                continue
            it = _item_to_int(row[idx_item])
            pu = to_number(row[idx_price])
            if it is None or pu is None:
                continue
            out[it] = pu  # si se repite, el último gana
        return out


def _leer_reporte_xlsx(path: Path) -> dict:
    wb = load_workbook(path, read_only=True, data_only=True)
    out = {}
    try:
        for ws in wb.worksheets:
            # buscar encabezado en primeras 80 filas
            header_row = None
            header_vals = None
            for r in range(1, min(ws.max_row, 80) + 1):
                vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
                vals_norm = [normalizar(v) for v in vals if v is not None]
                if not vals_norm:
                    continue
                if any("precio" in v for v in vals_norm) and any(v in ("numero", "nro", "item", "ítem") or "item" in v for v in vals_norm):
                    header_row = r
                    header_vals = vals
                    break

            if header_row is None or header_vals is None:
                continue

            headers_norm = [normalizar(v) if v is not None else "" for v in header_vals]
            item_h, price_h = _detectar_campos(headers_norm)
            if item_h is None or price_h is None:
                continue

            idx_item = headers_norm.index(item_h) + 1
            idx_price = headers_norm.index(price_h) + 1

            for r in range(header_row + 1, ws.max_row + 1):
                it = _item_to_int(ws.cell(row=r, column=idx_item).value)
                pu = to_number(ws.cell(row=r, column=idx_price).value)
                if it is None or pu is None:
                    continue
                out[it] = pu
    finally:
        wb.close()
    if not out:
        raise ValueError("No se encontraron filas válidas en el reporte (ítem + precio unitario).")
    return out


def parse_items_manual(texto: str) -> list:
    """
    Convierte '3,11-15,18' en [3,11,12,13,14,15,18]
    """
    if texto is None:
        return []
    s = normalizar(texto)
    s = s.replace(" ", "")
    if s == "":
        return []

    items = set()
    for part in s.split(","):
        if not part:
            continue
        if "-" in part:
            a, b = part.split("-", 1)
            ia = _item_to_int(a)
            ib = _item_to_int(b)
            if ia is None or ib is None:
                continue
            if ia > ib:
                ia, ib = ib, ia
            for x in range(ia, ib + 1):
                items.add(x)
        else:
            it = _item_to_int(part)
            if it is not None:
                items.add(it)

    return sorted(items)
