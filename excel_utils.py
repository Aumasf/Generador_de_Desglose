from openpyxl import load_workbook
import unicodedata

def normalizar(texto):
    """
    Quita acentos, pasa a minúsculas y elimina espacios extra.
    """
    texto = str(texto).strip().lower()
    texto = unicodedata.normalize("NFD", texto)
    texto = "".join(c for c in texto if unicodedata.category(c) != "Mn")
    return texto


def _es_item_valido(valor_item):
    """
    Retorna True solo si el item "existe" de verdad.
    Esto evita que se impriman líneas finales tipo firma/nombre.

    Acepta:
      - int: 1, 2, 3
      - float entero: 1.0, 2.0
      - str numérico: "1", "12", "101"
    Rechaza:
      - None
      - "" / "   "
      - strings no numéricos ("firma", "abc", etc.)
    """
    if valor_item is None:
        return False

    # string
    if isinstance(valor_item, str):
        txt = valor_item.strip()
        if txt == "":
            return False
        return txt.isdigit()

    # int
    if isinstance(valor_item, int):
        return True

    # float (si es entero tipo 1.0)
    if isinstance(valor_item, float):
        return valor_item.is_integer()

    return False


def _to_number(valor):
    """
    Convierte valores de Excel a número (float) de forma tolerante.

    - Excel suele venir como int/float -> se devuelve tal cual.
    - Si viene como string, intenta limpiar separadores de miles.
    """
    if valor is None:
        return None

    if isinstance(valor, (int, float)):
        return float(valor)

    if isinstance(valor, str):
        txt = valor.strip()
        if txt == "":
            return None

        # Limpieza típica (Gs.): "1.234.567" -> "1234567"
        # Si hubiese coma decimal, la convertimos a punto.
        txt = txt.replace(" ", "")
        txt = txt.replace(".", "")
        txt = txt.replace(",", ".")
        try:
            return float(txt)
        except ValueError:
            return None

    return None


def _buscar_titulo_llamado(ws, filas_busqueda=5):
    """
    Busca el título del llamado en las primeras filas.

    Regla:
    - La primera fila suele tener una celda que empieza con "Ítems del llamado" (o similar).
    - Ese texto debe ir completo debajo de la fecha en el PDF.

    Si no se encuentra un match claro, hace fallback a A1.
    """
    for fila in range(1, filas_busqueda + 1):
        for col in range(1, ws.max_column + 1):
            valor = ws.cell(row=fila, column=col).value
            if not valor:
                continue

            norm = normalizar(valor)
            if "items del llamado" in norm or "item del llamado" in norm:
                return str(valor).strip()

    v = ws.cell(row=1, column=1).value
    return str(v).strip() if v else ""


def _buscar_texto_lote(ws, filas_busqueda=15):
    """Busca un texto que contenga/empiece con "Lote" en las primeras filas.

    - Recorre las primeras *filas_busqueda* filas y todas las columnas.
    - Devuelve el primer match encontrado.
    - Si no encuentra nada, devuelve "".

    Nota:
    En varios Excels el contenido suele empezar con "Lote ...".
    """
    for fila in range(1, filas_busqueda + 1):
        for col in range(1, ws.max_column + 1):
            valor = ws.cell(row=fila, column=col).value
            if not valor:
                continue

            norm = normalizar(valor)

            # match fuerte: empieza con "lote" o "grupo"
            if norm.startswith("lote") or norm.startswith("grupo"):
                return str(valor).strip()

            # match flexible: contiene "lote" o "grupo"
            if "lote" in norm or "grupo" in norm:
                return str(valor).strip()


    return ""


def leer_items_y_descripciones_excel(ruta_excel):
    """
    Lee el Excel subido por el usuario y devuelve:

      (titulo_llamado, texto_lote, filas)

    filas: lista de dicts con:
        - item
        - descripcion
        - atributos (si existe la columna "Atributos"/"Atributo")
        - unidad_medida
        - presentacion
        - cantidad
        - precio_unitario_iva_incl
        - precio_total_iva_incl   (IMPORTANTE: es el TOTAL, no el unitario)
    """
    wb = load_workbook(ruta_excel, data_only=True)
    ws = wb.active

    # -------------------------------
    # Encabezados generales (parte superior del Excel)
    # -------------------------------
    titulo_llamado = _buscar_titulo_llamado(ws)
    texto_lote = _buscar_texto_lote(ws)

    fila_encabezados = None
    col_desc = None
    col_item = None
    col_unidad = None
    col_presentacion = None
    col_atributos = None
    col_cantidad = None
    col_precio_unit = None
    col_precio_total = None

    # Buscar encabezados en las primeras 10 filas
    for fila in range(1, 11):
        posibles = {}
        for col in range(1, ws.max_column + 1):
            valor = ws.cell(row=fila, column=col).value
            if valor:
                posibles[normalizar(valor)] = col

        # localizar descripción
        for nombre, col in posibles.items():
            if "descripcion" in nombre:
                fila_encabezados = fila
                col_desc = col
                break

        if fila_encabezados:
            # localizar Atributos (si existe) / Cantidad / Precios
            for nombre, col in posibles.items():
                if col_atributos is None and "atributo" in nombre:
                    col_atributos = col

            # Si existe 'Atributos', NO buscamos Unidad de Medida ni Presentación.
            if col_atributos is None:
                for nombre, col in posibles.items():
                    if col_unidad is None and "unidad" in nombre and "medida" in nombre:
                        col_unidad = col
                    if col_presentacion is None and "presentacion" in nombre:
                        col_presentacion = col

            # Cantidad / Precios (siempre)
            for nombre, col in posibles.items():
                if col_cantidad is None and nombre == "cantidad":
                    col_cantidad = col
                if col_precio_unit is None and "precio" in nombre and "unit" in nombre:
                    col_precio_unit = col
                if col_precio_total is None and "precio" in nombre and "total" in nombre:
                    col_precio_total = col

            break

    if not fila_encabezados:
        raise ValueError("No se encontró una fila de encabezados con una columna 'Descripción'")

    # localizar columna de item
    for col in range(1, ws.max_column + 1):
        valor = ws.cell(row=fila_encabezados, column=col).value
        if not valor:
            continue

        nombre = normalizar(valor)
        if nombre == "item" or nombre == "ítem" or nombre.startswith("item"):
            col_item = col
            break

    if not col_item:
        raise ValueError("No se encontró la columna 'Ítem' en el encabezado")

    filas = []
    for fila in range(fila_encabezados + 1, ws.max_row + 1):
        item = ws.cell(row=fila, column=col_item).value
        desc = ws.cell(row=fila, column=col_desc).value

        if not _es_item_valido(item):
            continue

        if desc is None or str(desc).strip() == "":
            continue

        atributos = ws.cell(row=fila, column=col_atributos).value if col_atributos else None
        unidad = ws.cell(row=fila, column=col_unidad).value if col_unidad else None
        presentacion = ws.cell(row=fila, column=col_presentacion).value if col_presentacion else None
        cantidad = ws.cell(row=fila, column=col_cantidad).value if col_cantidad else None
        precio_unit = ws.cell(row=fila, column=col_precio_unit).value if col_precio_unit else None
        precio_total = ws.cell(row=fila, column=col_precio_total).value if col_precio_total else None

        filas.append({
            "item": item,
            "descripcion": str(desc).strip(),
            # Si existe la columna 'Atributos' (o 'Atributo'), usamos ese texto y
            # dejamos vacíos Unidad/Presentación para evitar inconsistencias.
            "atributos": str(atributos).strip() if atributos is not None else "",
            "unidad_medida": "" if col_atributos else (str(unidad).strip() if unidad is not None else ""),
            "presentacion": "" if col_atributos else (str(presentacion).strip() if presentacion is not None else ""),
            "cantidad": _to_number(cantidad),
            "precio_unitario_iva_incl": _to_number(precio_unit),
            "precio_total_iva_incl": _to_number(precio_total),
        })

    return titulo_llamado, texto_lote, filas
